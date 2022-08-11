# 书写人：胡高远
# 书写时间：2022/7/15  16:00
import openpyxl
workbook = openpyxl.load_workbook('Excel.xlsx')
print(workbook.sheetnames)

# workbook.create_sheet('mysheet')

# sheet = workbook['mysheet']
# workbook.remove('sheet')

# sheet = workbook['产品信息']
# workbook.copy_worksheet(sheet)

sheet = workbook['产品信息 Copy']
sheet.title = 'mysheet'

print(workbook.sheetnames)

workbook.save('Excel.xlsx')


