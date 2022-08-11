# 书写人：胡高远
# 书写时间：2022/7/16  16:26
import openpyxl
from openpyxl.styles import Font,PatternFill
import re
workbook = openpyxl.load_workbook('../Excel.xlsx')
# print(workbook.sheetnames)
font = Font(color='ff0000')
fill = PatternFill(fill_type='solid',fgColor='ffff00')
sheet = workbook['mysheet2']
workbook.copy_worksheet(sheet)
sheet.title = '美国数据'
for row in workbook['美国数据'].rows:
    if re.search('.*美国.*',row[4].value):
        for i in range(0,10):
            row[i].fill = fill
            row[i].font = font
workbook.save('new Excel.xlsx')


