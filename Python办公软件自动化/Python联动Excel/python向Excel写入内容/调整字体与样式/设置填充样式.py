# 书写人：胡高远
# 书写时间：2022/7/16  15:53
import openpyxl
from openpyxl.styles import PatternFill   #填充
from openpyxl.styles import GradientFill   #渐变
workbook = openpyxl.load_workbook('project.xlsx')
sheet = workbook['Sheet1']
cell_c5 = sheet['c5']
pattern_fill = PatternFill(fill_type='solid',fgColor='ffff00')
cell_c5.fill = pattern_fill

cell_b1 = sheet['B1']
gradien = GradientFill(stop=('ff00ff','00ff00'))
cell_b1.fill = gradien

workbook.save('project.xlsx')
