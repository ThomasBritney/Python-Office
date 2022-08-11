# 书写人：胡高远
# 书写时间：2022/7/17  13:18
import openpyxl
from openpyxl.drawing.image import Image
workbook = openpyxl.load_workbook('new Excel.xlsx')
sheet = workbook.create_sheet('imagesheet')
# print(workbook.sheetnames)
# 创建图片对象
logo = Image('logo.png')
logo.height = 100
logo.width = 120
sheet.add_image(logo,'A1')
workbook.save('new Excel.xlsx')