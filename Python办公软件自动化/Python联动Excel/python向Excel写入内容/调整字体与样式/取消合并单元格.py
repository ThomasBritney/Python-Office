# 书写人：胡高远
# 书写时间：2022/7/16  16:13
import openpyxl

workbook = openpyxl.load_workbook('project.xlsx')
sheet = workbook['Sheet1']

# sheet.unmerge_cells('D1:G2')
sheet.unmerge_cells(start_row=1,start_column=10,end_row=10,end_column=15)
workbook.save('project.xlsx')